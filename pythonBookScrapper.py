# Auther: Luke Smith
# Description: It scrapes data from https://books.toscrape.com and saves it into several file formats(only excel right now).

""" 
Improvements:
    5. Loging instead of prints

    7. CSV andn JSON exports as well

    8. Add Retrying, timeouts, and skiping - so that one page failing to load doesn't end 
    the whole program
    9. Figure out user angents so that sites don't block the scrape
    10. Rate limits - sending a but-ton of requests too quickly will make the target site raise flags

    12. User request session
    13. Make congig folder

    Lastly. Once the scraper technicaly works, figure out all the "good product" stuff.
        Lastly. How would a customer run it? And what output would they actualy see?
"""

import requests                 # For getting html data from sites
from bs4 import BeautifulSoup   # For formating the html data in a way that is nice to work with
from openpyxl import Workbook   # For working with excel
from urllib.parse import urljoin# Has some functions to make working with urls easy
from book import Book           # For storing data on the books
import os                       # For handling filesystem stuff

print("Starting Program")

# getSoup - Gets the soup for the given url
# Parameters:
#       string url - The url from which a soup will be got. I am a poet
# Returns the soup - There are a TON of exceptions in here, so if this method runs successfully, you can trust it returns a functional soup
def getSoup(url):
    try:
        response = requests.get(url)
    except requests.exceptions.MissingSchema:
        print("Invalid URL (missing schema, like http://)")
        print("Given URL: " + url)
        raise Exception("Request Failure")
    except requests.exceptions.InvalidURL:
        print("Invalid URL format")
        print("Given URL: " + url)
        raise Exception("Request Failure")
    except requests.exceptions.ConnectionError:
        print("Failed to connect to server")
        print("Given URL: " + url)
        raise Exception("Request Failure")
    except requests.exceptions.Timeout:
        print("Request timed out")
        print("Given URL: " + url)
        raise Exception("Request Failure")
    except requests.exceptions.RequestException as e:
        print("Other request error:", e)
        print("Given URL: " + url)
        raise Exception("Request Failure: " + e)

    response.raise_for_status() # Raises an exception if it gets an unexpected status code like "404 Not Found"

    response.encoding = "utf-8"
    try:
        soup = BeautifulSoup(response.text, "html.parser")
    except Exception as e:
        raise Exception("Soup Failed: " + e)

    return soup

# incrementPageUrl - changes the pageUrl based on the new pageNum
# Parameters:
#       currentUrl - The url we are currently on
# Returns string - The currentUrl with the given page num in it.
def incrementPageUrl(currentUrl, soup):
    # find the link in the next button on the page. 
    nextButton = soup.find("li", class_="next")
    
    if (nextButton):
        nextPage = nextButton.find("a")["href"]
        nextUrl = urljoin(currentUrl, nextPage)
        return nextUrl
    else:
        raise Exception("Could not find Next Button with the next page URL")


# getNumberOfPages - Gets the number of pages of books in the website
# Parameters:
#       soup - The soup of the page it will search for the page num in
# Returns int - The number of pages. None if it failed to get a number
def getNumberOfPages(soup):
    # This will be the text in the <li> tag that has the page number
    pageOfText = soup.find("ul", class_="pager")

    if (pageOfText): # Checking that it found the <ul class="pager">
        pageOfText = pageOfText.find("li", class_="current")
        if (pageOfText): # Checking that it found the <li class="current">
            pageOfText = pageOfText.text
        else:
            return None # Return none of it could not find pageOfText
    else:
        return None # Return none of it could not find pageOfText
    
    # Assuming it found pageOfText, then we can start parsing it for the page number
    # The pageOfText should have something like "Page 1 of 50" or something.

    try:
        numOfPages = int(pageOfText.split()[-1])
        # This uses negative indexing, so yes, we are looking of the -1 index of the sequence.
        # That should be the last "word" in the pageOfText string, which should be the number of pages.
    except:
        return None
    
    if (numOfPages != -1):
        return numOfPages
    return None

# printBooks - Prints all the books in the given array of books
# Paremeters:
#       bookObjs - Array of books to print
# Returns nothing
def printBooks(bookObjs):
    for book in bookObjs:
        print("----------------------")
        print("Title: " + book.title)
        print("Price: " + book.price)
    print("----------------------")

# getBooksFromPage - Formats all the books in the given URL into book objects
# Paremeters:
#       soup - The soup of the page to be searched for books
# Returns: Array of books found at the given URL. Returns -1 if it failed. Returns none if no books were found.
def getBooksFromPage(soup):
    # Find all the books
    booksHtml = soup.find_all("article", class_="product_pod")

    # Books array:
    bookObjs = []

    for bookHtml in booksHtml:

        title_tag = bookHtml.select_one("h3 a")
        price_tag = bookHtml.select_one("p.price_color")

        # Checking that there is a title tag and price tag
        if (title_tag and price_tag):
            title = title_tag.get("title")
            # Note: tag["title"] would crash if there is no title, while .get("title") would return None

            # Checking that there is a title in the title_tag
            if (title):
                # Finaly, actualy puting the book data into a book object and into the bookObjs array
                bookObjs.append(
                    Book(
                        title,
                        price_tag.text
                    )
                )
    
    
    if (len(bookObjs) <= 0):
        return None
    
    return bookObjs

# makeWorkBookSheet - Compiles all the given book objects into an excel sheet
# Parameters:
#       bookObjs - Array of all the book objects to put into the exvel sheet
#       pageNum - The page number the books are from, and the page this excel sheet will be
#       sheet - The sheet object that the bookObjs will be put into.
# Returns: Nothing
def makeWorkBookSheet(bookObjs, pageNum, sheet):
    # Making the sheet:
    sheet.title = str(pageNum) + "pages of books"

    sheet.append(["", ""])
    sheet.append(["Page: " + str(pageNum), ""])
    sheet.append(["", ""])

    # Making a header:
    sheet.append(["Title", "Price"])

    # Looping through each book:
    for book in bookObjs:
        sheet.append(book.getRowOfData())

# getUserInput - Gets the user input for how many pages the program will look through
# Parameters: None
# Returns: The number of pages the user wants to scrape from. -1 If the user wants to quit 
def getUserInput(pageUrl):
    soup = None
    try:
        soup = getSoup(pageUrl)
    except Exception as e:
        raise Exception("Soup Failed: " + e)

    numOfPages = getNumberOfPages(soup)

    if (numOfPages == None):
        raise Exception("Failed to get Pages: Returned None.")
    if (numOfPages < 0):
        raise Exception("Failed to get Pages: Out of range.")

    print("Total Number of pages: " + str(numOfPages))

    userInput = -1
    while (userInput < 0 or userInput > numOfPages):
        print("\nInput 0 to leave the program")
        userInput = input("How many pages do you want to scrape?: ")
        try:
            userInput = int(userInput)
            if (userInput < 0 or userInput > numOfPages):
                print("That number is out of range. Please try again")
        except:
            print("That is not a valid number. Please try again")
            userInput = -1 # Setting it to -1 so the while loop starts over

    return userInput

# savePageToWorkbook - Saves the given page to the given workbook
# Parameters:
#       workbook - The workbook which the page will be saved to
#       page - the page to be saved
#       pageNumber - The page number that we are on
# Returns: Nothing directly, just modifies the workbook
def savePageToWorkbook(workbook, page, pageNumber):
    sheet = workbook.active

    makeWorkBookSheet(page, pageNumber, sheet)

# - - - [ Variables:  ] - - - 
pageUrl = "https://books.toscrape.com"  # The URL of the page we are currently on
numOfPages = getUserInput(pageUrl)      # Gets user input for the number of pages to scrape data from
workBook = Workbook()                   # Workbook for saving the data into an excel doc
pageNum = 1                             # The page number of the page we are currently on
# - - - - - - - - - - - - - - 

# -- -- --[  Constants:  ]-- -- --
lengthOfBar = 30                        # The length of the loading bar
outputFileName = "outPutFile"           # The name of the ouputfiles
# -- -- -- -- -- -- -- -- -- -- --


if (numOfPages != 0):
    print("Scraping from pages.")
    print("|" + "-"*lengthOfBar + "|  (0 / " + str(numOfPages) + ")", end="\r")

    # Loop through each page:
    for index in range(0, numOfPages):
        # Get the soup of the page:
        try:
            soup = getSoup(pageUrl)
        except Exception as e:
            print("Soup encountered exception: " + e)
            continue # Move on to next page if the soup fails
            # ToDo: Add Retry logic here

        # Get the books from the current page
        thisPage = getBooksFromPage(soup)

        # If we got stuff from this page:
        if (thisPage != None):
            # Then push the books from this page to the pages array of arrays:
            #pages.append(thisPage)

            savePageToWorkbook(workBook, thisPage, pageNum)

            # Now the url is incremented
            try:
                pageUrl = incrementPageUrl(pageUrl, soup)
            except Exception as e:
                print("Failed to get next page url")
                print("Error: " + e)
                break
                # ToDo: Add retry logic to refresh the page/soup and try incrementPageUrl() again.

            pageNum += 1

            # And we update the loading bar:
            numOfEquals = int((lengthOfBar/numOfPages) * (index + 1))
            numOfDashes = lengthOfBar - numOfEquals
            print("|" + "="*numOfEquals + "-"*numOfDashes + "|  (" + str(index + 1) + " / " + str(numOfPages) + ")", end="\r")
        else:
            print("Failed to get books from page " + str(index + 1) + ".")

    print("\n")

# Save to an excel document
folderPath = "~/WindowsSucks"   # Linux/WSL file path
folderPath = os.path.expanduser(folderPath)

# Create folder if it doesn't exist
os.makedirs(folderPath, exist_ok=True)

#excelDocName = folderPath + outputFileName + ".xlsx"
excelDocName = os.path.join(folderPath, outputFileName + ".xlsx")

print("Saving to: " + excelDocName)
workBook.save(excelDocName)

print("Thank you for using this program!")